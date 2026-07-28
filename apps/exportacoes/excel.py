"""
Exportação da grade para Excel (Módulo 14), usando openpyxl.

Recebe a mesma estrutura de grid já usada nas telas de Grade (Módulo 10)
e Relatórios (Módulo 13) — {horario: {codigo_dia: aula_ou_None}} — e
`montar_linhas(aula)`, uma função que decide o que aparece em cada célula
(a Grade por turma mostra disciplina/professor/ambiente; o relatório por
professor mostra turma/disciplina/ambiente).
"""
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

AZUL_CABECALHO = '0D6EFD'


def gerar_excel_grade(titulo, dias, grid, montar_linhas):
    workbook = Workbook()
    planilha = workbook.active
    planilha.title = 'Grade'

    total_colunas = len(dias) + 1
    planilha.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_colunas)
    celula_titulo = planilha.cell(row=1, column=1, value=titulo)
    celula_titulo.font = Font(bold=True, size=14)
    celula_titulo.alignment = Alignment(horizontal='center')

    preenchimento = PatternFill(start_color=AZUL_CABECALHO, end_color=AZUL_CABECALHO, fill_type='solid')
    linha_cabecalho = 3
    celula = planilha.cell(row=linha_cabecalho, column=1, value='Horário')
    celula.font = Font(bold=True, color='FFFFFF')
    celula.fill = preenchimento
    celula.alignment = Alignment(horizontal='center')
    for indice, (_codigo, rotulo) in enumerate(dias, start=2):
        celula = planilha.cell(row=linha_cabecalho, column=indice, value=rotulo)
        celula.font = Font(bold=True, color='FFFFFF')
        celula.fill = preenchimento
        celula.alignment = Alignment(horizontal='center')

    linha_atual = linha_cabecalho + 1
    for horario, por_dia in grid.items():
        planilha.cell(
            row=linha_atual, column=1,
            value=f'{horario.inicio.strftime("%H:%M")}–{horario.fim.strftime("%H:%M")}',
        ).alignment = Alignment(horizontal='center', vertical='center')
        for indice, (codigo, _rotulo) in enumerate(dias, start=2):
            aula = por_dia.get(codigo)
            texto = '\n'.join(montar_linhas(aula)) if aula else ''
            celula = planilha.cell(row=linha_atual, column=indice, value=texto)
            celula.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        linha_atual += 1

    for coluna in range(1, total_colunas + 1):
        planilha.column_dimensions[get_column_letter(coluna)].width = 22
    for linha in range(linha_cabecalho, linha_atual):
        planilha.row_dimensions[linha].height = 45

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer
